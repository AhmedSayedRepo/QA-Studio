"""task_manager.py — the "Task Manager" screen.

Two independent tools, both scoped to a sprint/iteration (matching the app's
existing convention — see Sprint Plan / Sprint Report — rather than a raw date
range, since Original Estimate / Completed Work aren't themselves timestamped
fields in Azure DevOps):

  1. Task workload report — pick a sprint + a project member, and see every
     'Task' work item assigned to them in that sprint, with Original Estimate
     and Completed Work summed (plus a per-task breakdown table).
  2. Bulk child-task creation — pick the same sprint, select one or more User
     Stories in it, give each one its own child-Task title, and assign every
     new task to one chosen member in a single run.

Self-contained: own `_tm_*` state, reuses regression.py's checkbox multiselect
/ id-link / avatar helpers and engine.py's Azure calls. main.py imports this
module and dispatches `screen(app)` for the "task_manager" nav tab.
"""
import os
import re
import json
import threading
from datetime import date, datetime, timedelta

import flet as ft
import theme as T
import engine as E

# A story can get up to this many child tasks in one "Create child tasks" run
# (a batch/patch of tasks per story) — keeps a single run bounded and the UI
# from growing unboundedly if someone clicks "Add another task" repeatedly.
_TM_CT_MAX_ROWS = 10

# How long Calculate waits for Azure DevOps before giving up and surfacing a
# timeout error, rather than leaving the button stuck on "Calculating…"
# forever if the request itself hangs (not a Python exception — see the
# watchdog note in _calc_report). Module-level so a test can shrink it.
_TM_REPORT_TIMEOUT_S = 90

# The org's standard full-time workload: 170 hours in an average month —
# used as a separate "how does Completed Work compare to a normal month's
# capacity" benchmark, alongside (never replacing) the existing Original
# Estimate-based Remaining figure, for BOTH scope modes:
#   - Date range mode: prorated over the picked start/end dates.
#   - Sprint mode: prorated over the SPRINT's own start/finish dates (from
#     Azure DevOps' iteration data — see engine.fetch_iterations), not the
#     count of tasks in it, since a sprint can run short or long.
# Calendar days (not business/working days) — a fixed, explicit choice, since
# 170h/month is itself an average-month figure, not tied to any specific
# month's actual weekday count.
_TM_MONTHLY_BENCHMARK_HOURS = 170.0
_TM_AVG_DAYS_PER_MONTH = 30.44   # 365.25 / 12 — the standard average month length


def _tm_benchmark_hours(days):
    """Prorated share of the 170h/month benchmark for a `days`-long period
    (inclusive of both endpoints — see _tm_period_days). Returns None if
    `days` is unknown/non-positive (e.g. a sprint with no dates configured
    in Azure DevOps), so callers can show "—" instead of a misleading 0."""
    if not days or days <= 0:
        return None
    return _TM_MONTHLY_BENCHMARK_HOURS / _TM_AVG_DAYS_PER_MONTH * days


def _tm_fmt_date(s):
    """"YYYY-MM-DD" -> "Jul 1, 2026" for the benchmark-period caption below
    the KPI strip — falls back to the raw string (or an em dash) rather than
    raising if it's missing/malformed. Built from `d.strftime("%b")` plus
    `d.day`/`d.year` interpolated directly (rather than strftime's "%-d"/
    "%#d" no-leading-zero flag, which is spelled differently on Linux/macOS
    vs Windows glibc — this app runs on Windows, where "%-d" raises) — no
    leading zero ever appears in the first place, so there's nothing fragile
    to strip back out afterward."""
    try:
        if not s:
            return "—"
        d = date.fromisoformat(s)
        return f"{d.strftime('%b')} {d.day}, {d.year}"
    except Exception:
        return s or "—"


def _tm_period_days(start_str, end_str):
    """Inclusive calendar-day count between two "YYYY-MM-DD" strings (a
    sprint that runs Mon-Fri of one week is 5 days IN THE SENSE OF ITS OWN
    date span here — both start and end date count, matching how a "Sprint
    lasts 14 days" is normally meant). Returns None if either date is
    missing/unparseable rather than raising, since callers treat that as
    "benchmark not available" instead of a hard error."""
    try:
        if not (start_str and end_str):
            return None
        return (date.fromisoformat(end_str) - date.fromisoformat(start_str)).days + 1
    except Exception:
        return None


def _tm_default_row(st):
    """A fresh, empty per-task row for the bulk child-task creation section.
    `st` is the story dict (for the default "Testing - <story title>" title)."""
    return {"title": f"Testing - {st.get('title', '')}", "due": "",
            "estimate": "", "completed": ""}


def _tm_display_name(app, email):
    """Resolve a project member's email to their display name, via the same
    member cache the 'Assigned to' dropdown is built from — falls back to the
    raw email if the cache hasn't loaded yet or the address isn't a match.
    Used everywhere the workload report shows WHO it's for (exports, email,
    filenames), so a Task Manager report reads like "Ahmed Sayed" instead of
    "ahmed@example.com", the way the rest of the app's reports do."""
    email = (email or "").strip()
    for m in (getattr(app, "_members_cache", None) or []):
        if (m.get("email") or "").lower() == email.lower():
            return m.get("name") or email
    return email


def _init(app):
    for k, v in (
        ("_tm_iterations", []), ("_tm_iter_loading", False), ("_tm_iter_for", None),
        ("_tm_iteration", ""),
        # Explicit toggle between the two ways to scope Calculate (Section 2)
        # — "sprint" or "dates". Only the active mode's value is used; the
        # inactive one is visibly de-emphasized in the UI (see screen()) but
        # its state is kept, not wiped, so flipping back and forth doesn't
        # lose what was typed. Dates are two "YYYY-MM-DD" strings, same
        # convention as ai_usage_screen's range.
        ("_tm_scope_mode", "sprint"),
        # Bumped once at the top of every screen() call (a FULL rebuild),
        # captured by that call's _sync_report_dynamic/_sync_ct_dynamic
        # closures as their "epoch". See the note beside the bump in
        # screen() for why this exists — it lets a closure from an older,
        # superseded screen() notice it's stale and fall back to a full
        # app.render() instead of silently updating a dead, detached cell.
        ("_tm_screen_epoch", 0),
        ("_tm_start_date", ""), ("_tm_end_date", ""),
        # workload-report section
        ("_tm_report_user", ""), ("_tm_report_busy", False),
        # Bumped by every _calc_report call and captured by that call's
        # background thread — a fetch whose generation no longer matches
        # app._tm_report_gen by the time it finishes was superseded by a
        # later Calculate click (e.g. a slow org-wide date-range query still
        # in flight when the user switches to Sprint mode and clicks
        # Calculate again) and must not touch busy/result/msg at all, so a
        # stale completion can never clobber a newer one or leave the button
        # stuck showing "Calculating…" forever.
        ("_tm_report_gen", 0),
        ("_tm_report_result", None), ("_tm_report_msg", None),
        ("_tm_report_export_msg", None),
        ("_tm_report_email_to", ""), ("_tm_report_email_open", False),
        ("_tm_report_emailing", False),
        # bulk child-task section
        ("_tm_ct_stories", []), ("_tm_ct_stories_loading", False),
        ("_tm_ct_stories_for", None),
        ("_tm_ct_selected", []), ("_tm_ct_open", False),
        # _tm_ct_rows: {story_id: [{"title","due","estimate","completed"}, ...]}
        # — a LIST per story so each selected story can get a batch/patch of
        # up to _TM_CT_MAX_ROWS child tasks in one run, not just one.
        ("_tm_ct_rows", {}), ("_tm_ct_user", ""),
        ("_tm_ct_busy", False), ("_tm_ct_result", None), ("_tm_ct_msg", None),
    ):
        if not hasattr(app, k):
            setattr(app, k, v)


def _date_field(app, label, value_str, on_pick, on_after=None, disabled=False, min_date=None):
    """Click-to-open calendar field backed by ft.DatePicker (Flet's native
    date overlay) — same pattern as ai_usage_screen._date_field (plain
    caption label above the box, not a floating Material label, which was
    getting clipped on dense fields). Unlike the AI Usage report's date
    range (which can never exceed today), a Task's Due Date is a future
    deadline, so this allows several years ahead and has no lower/upper
    bound tied to "today" beyond a sane calendar window.
    `disabled=True` mutes the box and stops it from opening the picker —
    used by the Task Manager sprint/date-range toggle so the inactive side
    reads as visibly inert instead of just as clickable as the active one.
    `min_date` ("YYYY-MM-DD"), if given, becomes the calendar's earliest
    selectable day — used to stop End date from ever being picked before
    Start date, rather than only catching it after the fact."""
    try:
        val = date.fromisoformat(value_str) if value_str else date.today()
    except Exception:
        val = date.today()

    try:
        first = date.fromisoformat(min_date) if min_date else date(2020, 1, 1)
    except Exception:
        first = date(2020, 1, 1)
    if val < first:
        val = first

    ink = T.INK_3 if disabled else (T.INK if value_str else T.INK_3)
    value_text = ft.Text(value_str or "Pick a date", size=12.5, weight=ft.FontWeight.W_600,
                         color=ink, font_family=T.F_MONO)
    box = ft.Container(
        ft.Row([ft.Icon(ft.Icons.CALENDAR_MONTH_OUTLINED, size=15,
                       color=T.INK_3), value_text], spacing=8),
        width=150, height=40, padding=ft.Padding.symmetric(horizontal=12),
        alignment=ft.Alignment.CENTER_LEFT,
        border=ft.Border.all(1, T.BORDER), border_radius=T.R,
        bgcolor=(T.BORDER_2 if disabled else T.CARD_2), opacity=(0.55 if disabled else 1.0))

    def _changed(e):
        d = e.control.value
        if d:
            # Same off-by-one DatePicker quirk worked around in
            # ai_usage_screen._date_field (flet-dev/flet#6145, still open as
            # of Flet 0.85.3): on_change reports one day earlier than tapped.
            d = d + timedelta(days=1)
            on_pick(d.strftime("%Y-%m-%d"))
        if on_after:
            on_after()
        else:
            app.ui_safe(app.render)

    dp = ft.DatePicker(value=val, first_date=first,
                       last_date=date(date.today().year + 5, 12, 31), on_change=_changed)

    def _open(e):
        try:
            app.page.overlay[:] = [c for c in app.page.overlay
                                   if not isinstance(c, ft.DatePicker)]
        except Exception:
            pass
        try:
            app.page.overlay.append(dp)
            dp.open = True
            app.page.update()
        except Exception:
            pass
        try:
            app.page.open(dp)
        except Exception:
            pass

    # GestureDetector raises at runtime if it has zero event handlers wired —
    # a disabled field has nothing to click, so skip wrapping it at all
    # rather than passing on_tap=None (which trips that requirement).
    clickable = box if disabled else ft.GestureDetector(content=box, on_tap=_open)
    return ft.Column([
        ft.Text(label, size=10.5, weight=ft.FontWeight.BOLD, color=T.INK_3),
        clickable,
    ], spacing=6, tight=True)


def _sprint_num(text):
    m = re.search(r"[Ss]print\s*\d+", text or "")
    return re.sub(r"\s+", " ", m.group(0)).strip() if m else ""


def _tm_scope_label(res):
    """Human-readable "which period" label for a workload-report result — a
    sprint number if Calculate was scoped to one, otherwise the date range
    that was used instead (Calculate now accepts either). Used everywhere the
    report shows/exports WHAT period it covers, so those spots don't each
    re-derive the sprint-vs-dates fallback themselves."""
    sp = _sprint_num(res.get("iteration", "")) or res.get("iteration", "")
    return sp or res.get("date_range", "") or "report"


def _sort_key(it):
    m = re.search(r"\d+", _sprint_num(it.get("name", "")) or _sprint_num(it.get("path", "")))
    return int(m.group(0)) if m else -1


def _load_iterations(app):
    if not (app.connected and app.project):
        return
    if app._tm_iter_loading or getattr(app, "_tm_iter_for", None) == app.project:
        return
    app._tm_iter_loading = True
    _proj = app.project
    app._tm_iter_for = _proj

    def _work():
        try:
            its = E.fetch_iterations(_proj) or []
        except Exception:
            its = []
        sprints = [it for it in its
                   if (_sprint_num(it.get("name", "")) or _sprint_num(it.get("path", "")))] or its
        sprints.sort(key=lambda it: (_sort_key(it) < 0, _sort_key(it)))
        app._tm_iter_loading = False
        if app.project != _proj:
            app._tm_iter_for = None
            return
        app._tm_iterations = sprints
        app.ui_safe(app.render)
    threading.Thread(target=_work, daemon=True).start()


def _load_members(app):
    if (getattr(app, "_members_cache", None) is None
            and not getattr(app, "_members_loading", False) and app.project):
        app._members_loading = True

        def _work():
            try:
                mem = E.fetch_project_members(app.project)
            except Exception:
                mem = []
            app._members_cache = mem
            app._members_loading = False
            app.ui_safe(app.render)
        threading.Thread(target=_work, daemon=True).start()


def _load_stories(app):
    """Stories for the child-task picker, for the currently selected sprint —
    cached per (project, iteration) pair (same guard shape as _load_iterations,
    so switching sprints doesn't re-fire while a fetch is already in flight)."""
    it = app._tm_iteration
    if not (app.connected and app.project and it):
        return
    key = (app.project, it)
    if app._tm_ct_stories_loading or getattr(app, "_tm_ct_stories_for", None) == key:
        return
    app._tm_ct_stories_loading = True
    app._tm_ct_stories_for = key

    def _work():
        try:
            rows = E.fetch_stories_in_iteration(app.project, it) or []
        except Exception:
            rows = []
        app._tm_ct_stories_loading = False
        if (app.project, app._tm_iteration) != key:
            app._tm_ct_stories_for = None
            return
        app._tm_ct_stories = rows
        app.ui_safe(app.render)
    threading.Thread(target=_work, daemon=True).start()


def _calc_report(app, on_update=None):
    """`on_update`, if given, replaces `app.render()` as the thing that gets
    run (still via `app.ui_safe`, so it's safe to call from the background
    thread below) after each state change — lets the Task Manager screen pass
    in a small targeted cell-rebuild instead of a full-page rebuild for every
    click of Calculate, so this button no longer forces the whole nav+screen
    tree to redraw just to show a spinner or a result."""
    _target = on_update or app.render

    def _upd():
        app.ui_safe(_target)

    if getattr(app, "readonly", False):
        app._tm_report_msg = ("err", "Your role doesn't allow using Task Manager.")
        _upd()
        return
    mode = getattr(app, "_tm_scope_mode", "sprint")
    it = app._tm_iteration
    start = (app._tm_start_date or "").strip()
    end = (app._tm_end_date or "").strip()
    user = app._tm_report_user
    # Calculate is scoped by the explicit Sprint/Date range toggle, not by
    # which fields merely happen to be filled in — so switching the toggle
    # to "Sprint" always uses the sprint, ignoring stale dates left over from
    # "Date range" mode, and vice versa.
    use_dates = (mode == "dates")
    if use_dates:
        it = ""
    else:
        start = end = ""
    if not it and not use_dates:
        app._tm_report_msg = ("err", "Pick a sprint first.")
        _upd()
        return
    if use_dates and not (start and end):
        app._tm_report_msg = ("err", "Pick a start and end date first.")
        _upd()
        return
    if use_dates and end < start:
        app._tm_report_msg = ("err", "End date must be on or after the start date.")
        _upd()
        return
    if not user:
        app._tm_report_msg = ("err", "Pick a person first.")
        _upd()
        return
    # A new generation for THIS click — captured by _work() below and checked
    # again when the fetch finishes. Switching Sprint<->Date range and
    # clicking Calculate again starts a brand-new fetch without waiting for
    # an earlier one to finish (the date-range branch below queries an
    # entire project's Task history with no server-side assignee filter, so
    # it can legitimately take a long time on an active project — the app
    # must not look "stuck" behind it just because an older, abandoned fetch
    # hasn't returned yet). Only the LATEST generation's completion is
    # allowed to touch busy/result/msg; anything older is discarded quietly.
    app._tm_report_gen = getattr(app, "_tm_report_gen", 0) + 1
    my_gen = app._tm_report_gen
    app._tm_report_busy = True
    app._tm_report_result = None
    app._tm_report_msg = None
    _upd()

    # Belt-and-suspenders alongside the generation guard above: that guard
    # only helps once SOME newer click supersedes a stuck one — it does
    # nothing if the fetch itself simply never returns (a genuine network
    # stall/hang against Azure DevOps, not a Python exception — confirmed via
    # diagnostics.log showing no exception at all logged for a real "stuck"
    # report, which rules out a silently-swallowed crash and points at the
    # request itself hanging). A `requests` call CAN still block forever
    # past its own `timeout=` in rare cases (DNS hangs, some proxy/firewall
    # states) since that timeout only bounds socket read/connect stalls, not
    # every failure mode. This watchdog guarantees the BUTTON recovers within
    # a bounded time regardless: if `_work()` hasn't finished by the
    # deadline, treat it exactly like a superseded click — bump the
    # generation (so the fetch's eventual, late completion is discarded by
    # the same check in `_work()`, never applies stale state) — and surface
    # a clear, actionable message instead of leaving "Calculating…" up
    # forever with no way to tell if it's working or frozen.
    def _give_up():
        if getattr(app, "_tm_report_gen", 0) != my_gen or not app._tm_report_busy:
            return   # already finished (or superseded) on its own
        app._tm_report_gen = my_gen + 1
        app._tm_report_busy = False
        app._tm_report_msg = ("err",
            "Azure DevOps didn't respond in time. Try again — if it's scoped to "
            "a date range, a narrower range or picking the sprint instead is "
            "usually faster.")
        _upd()

    watchdog = threading.Timer(_TM_REPORT_TIMEOUT_S, _give_up)
    watchdog.daemon = True
    watchdog.start()

    def _work():
        try:
            if it:
                res = E.fetch_user_task_stats(app.project, iteration_path=it, assignee=user)
            else:
                res = E.fetch_user_task_stats(app.project, assignee=user,
                                              start_date=start, end_date=end)
            res["assignee_name"] = _tm_display_name(app, res.get("assignee", ""))
            err = None
        except Exception as ex:
            res = None
            err = str(ex)[:160]
        finally:
            watchdog.cancel()   # fetch finished before the deadline — no-op if it already fired
        if getattr(app, "_tm_report_gen", 0) != my_gen:
            return   # superseded by a newer click, or the watchdog already gave up
        if res is not None:
            app._tm_report_result = res
            if res.get("count", 0) == 0:
                app._tm_report_msg = ("err", "No tasks assigned to this person in that period.")
        else:
            app._tm_report_msg = ("err", f"Couldn't load task stats: {err}")
        app._tm_report_busy = False
        _upd()
    threading.Thread(target=_work, daemon=True).start()


def _create_tasks(app):
    if getattr(app, "readonly", False):
        app._tm_ct_msg = ("err", "Your role doesn't allow using Task Manager.")
        app.ui_safe(app.render)
        return
    sel = list(app._tm_ct_selected or [])
    user = app._tm_ct_user
    if not sel:
        app._tm_ct_msg = ("err", "Select at least one user story first.")
        app.ui_safe(app.render)
        return
    if not user:
        app._tm_ct_msg = ("err", "Pick who the tasks should be assigned to.")
        app.ui_safe(app.render)
        return
    stories_by_id = {s["id"]: s for s in app._tm_ct_stories}
    items = []
    for sid in sel:
        st = stories_by_id.get(sid, {})
        rows = app._tm_ct_rows.get(sid) or [_tm_default_row(st)]
        for row in rows[:_TM_CT_MAX_ROWS]:
            title = (row.get("title") or f"Testing - {st.get('title', '')}").strip()
            if not title:
                continue
            items.append({
                "story_id": sid, "title": title, "assigned_to": user,
                "due_date": (row.get("due") or "").strip(),
                "original_estimate": row.get("estimate"),
                "completed_work": row.get("completed"),
            })
    if not items:
        app._tm_ct_msg = ("err", "Nothing to create — add at least one task with a title.")
        app.ui_safe(app.render)
        return

    app._tm_ct_busy = True
    app._tm_ct_result = None
    app._tm_ct_msg = None
    app.ui_safe(app.render)

    def _work():
        try:
            res = E.create_child_tasks(app.project, items)
            app._tm_ct_result = res
            if res.get("ok"):
                app._tm_ct_msg = ("ok", f"Created {res['ok']} task(s).")
            if res.get("errors"):
                app._tm_ct_msg = ("err", f"{len(res['errors'])} failed — see details below.")
        except Exception as ex:
            app._tm_ct_msg = ("err", f"Couldn't create tasks: {str(ex)[:160]}")
        app._tm_ct_busy = False
        app.ui_safe(app.render)
    threading.Thread(target=_work, daemon=True).start()


def _status_banner(msg, busy):
    if not msg or busy:
        return None
    kind, text = msg
    ok = (kind == "ok")
    return ft.Container(
        ft.Row([ft.Icon(ft.Icons.CHECK_CIRCLE if ok else ft.Icons.ERROR_OUTLINE,
                       color=(T.GREEN if ok else T.RED), size=18),
               ft.Text(text, color=(T.GREEN if ok else T.RED), size=12.5, expand=True)],
              spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER),
        padding=ft.Padding.symmetric(vertical=11, horizontal=14),
        margin=ft.Margin.only(top=14), border_radius=T.R,
        bgcolor=(T.GREEN_SOFT if ok else T.RED_SOFT))


def _tm_out_dir():
    d = os.path.join(os.path.expanduser("~"), "QA Studio", "Task Manager Reports")
    os.makedirs(d, exist_ok=True)
    return d


def _tm_stamp(app, res):
    """Shared filename base for every export (Excel/PDF/JSON) and email
    attachment of the workload report — includes "Workload" explicitly so
    the file reads as e.g. "MyProject_Workload_Sprint42_ahmed.xlsx", not just
    "MyProject_Sprint42_ahmed.xlsx", which on its own didn't say what kind of
    report it was once saved outside the app (next to other QA Studio
    exports in the same folder)."""
    scope = _tm_scope_label(res)
    who = (res.get("assignee_name") or res.get("assignee", "") or "").split("@")[0] or "report"
    base = f"{app.project}_Workload_{scope}_{who}"
    return re.sub(r"[^A-Za-z0-9_-]+", "_", base).strip("_") or "task_workload_report"


def _tm_export_json(app, res):
    p = os.path.join(_tm_out_dir(), _tm_stamp(app, res) + ".json")
    with open(p, "w", encoding="utf-8") as f:
        json.dump(res, f, ensure_ascii=False, indent=2)
    return p


def _tm_export_xlsx(app, res):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from datetime import datetime as _dt
    wb = Workbook()
    ws = wb.active
    ws.title = "Task Report"
    head = Font(bold=True, color="FFFFFF", name="Segoe UI")
    fill = PatternFill("solid", fgColor="3A57D6")
    thin = Border(*[Side(style="thin", color="E6E8F1")] * 4)
    _AR = Alignment(horizontal="right", vertical="center")
    r = 1
    for k, v in (("Project", app.project), ("Sprint", res.get("iteration", "")),
                 ("Assigned to", res.get("assignee_name") or res.get("assignee", "")),
                 ("Generated", _dt.now().strftime("%Y-%m-%d %H:%M"))):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
        r += 1
    r += 1
    cols = ["Task ID", "Title", "Story", "State", "Original Estimate (h)", "Completed Work (h)"]
    for c, name in enumerate(cols, 1):
        cell = ws.cell(r, c, name)
        cell.font = head
        cell.fill = fill
    r += 1
    for t in res.get("tasks", []):
        vals = [t["id"], t["title"], t.get("parent_title") or "", t["state"],
                t["original_estimate"], t["completed_work"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            if c in (1, 5, 6):
                cell.alignment = _AR
        r += 1
    ws.cell(r, 4, "TOTAL").font = Font(bold=True)
    ws.cell(r, 5, res.get("total_original_estimate", 0)).font = Font(bold=True)
    ws.cell(r, 6, res.get("total_completed_work", 0)).font = Font(bold=True)
    for c, w in zip("ABCDEF", [10, 46, 30, 14, 20, 20]):
        ws.column_dimensions[c].width = w
    p = os.path.join(_tm_out_dir(), _tm_stamp(app, res) + ".xlsx")
    wb.save(p)
    return p


def _tm_export_pdf(app, res):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import regression as R
    fn = R._pdf_font()
    styles = getSampleStyleSheet()
    for sn in ("Title", "Normal"):
        try:
            styles[sn].fontName = fn
        except Exception:
            pass
    p = os.path.join(_tm_out_dir(), _tm_stamp(app, res) + ".pdf")
    doc = SimpleDocTemplate(p, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm)
    elems = [Paragraph(R._ar("Task Workload Report"), styles["Title"]),
             Paragraph(R._ar(f"{app.project} · {res.get('iteration', '')} · "
                            f"{res.get('assignee_name') or res.get('assignee', '')}"), styles["Normal"]),
             Spacer(1, 8 * mm)]
    data = [["ID", "Title", "Story", "State", "Est.", "Completed"]]
    for t in res.get("tasks", []):
        data.append([str(t["id"]), R._ar((t["title"] or "")[:44]),
                    R._ar((t.get("parent_title") or "")[:28]), t["state"],
                    f"{t['original_estimate']:g}", f"{t['completed_work']:g}"])
    data.append(["", "", "", "TOTAL",
                f"{res.get('total_original_estimate', 0):g}",
                f"{res.get('total_completed_work', 0):g}"])
    tbl = Table(data, colWidths=[14 * mm, 62 * mm, 40 * mm, 20 * mm, 18 * mm, 22 * mm],
               repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3A57D6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), fn),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E6E8F1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F6F8FC")]),
        ("FONTNAME", (3, -1), (-1, -1), fn),
    ]))
    elems.append(tbl)
    doc.build(elems)
    return p


def _tm_report_html(app, res):
    """Polished, Outlook-safe HTML for the task workload report email — SAME
    masthead / hero pill / metric-strip / section-head / footer design
    language as regression.py's _plan_html (Sprint Plan / Regression Plan
    email) and engine.py's build_report_email / build_sprint_summary_email,
    so this is no longer a bare unstyled table — all the app's report emails
    now read as one consistent brand.

    Every free-text field below (task/story title, assignee display name,
    state) is passed through html.escape() before being dropped into this
    f-string HTML — those values come from Azure DevOps work items (titles
    anyone with edit access to the project could set to anything, including
    literal HTML/script markup) or the local member cache, not from a fixed
    set this app controls, so they can't be trusted to already be safe HTML.
    Same convention engine.py's own report emails already use throughout
    (build_report_email, build_sprint_summary_email, build_ai_usage_email)."""
    import html as _html
    import regression as R
    PAPER = "#E9E8EE"; CARD = "#FFFFFF"; TINT = "#FAFAFC"
    INK = "#1B1A22"; INK2 = "#6B6975"; INK3 = "#9C9AA6"
    LINE = "#E8E7EE"; LINE2 = "#F1F0F5"
    VIOLET = "#0E9CC0"; VIOLET_INK = "#0B6E86"; VIOLET_SOFT = "#D6F4FB"
    GREEN = "#1F8A52"
    UI = '"Segoe UI",Roboto,Helvetica,Arial,sans-serif'
    MONO = '"SFMono-Regular",Consolas,Menlo,monospace'

    def _sec_head(dot, title, count):
        return (f"<table role='presentation' cellpadding='0' cellspacing='0'><tr>"
                f"<td valign='middle' style='padding-right:10px'><span style='display:inline-block;"
                f"width:9px;height:9px;border-radius:50%;background:{dot}'></span></td>"
                f"<td valign='middle' style='font-size:14.5px;font-weight:700;color:{INK};"
                f"letter-spacing:-.2px'>{title}</td>"
                f"<td valign='middle' style='padding-left:9px'><span style='font-family:{MONO};"
                f"font-size:11px;font-weight:700;color:{INK2};background:{LINE2};border-radius:20px;"
                f"padding:3px 9px'>{count}</span></td></tr></table>")

    tasks = res.get("tasks", [])
    est_total = res.get("total_original_estimate", 0)
    comp_total = res.get("total_completed_work", 0)

    metrics_data = [("Tasks", res.get("count", 0), INK),
                    ("Original est.", f"{est_total:g}h", VIOLET_INK),
                    ("Completed", f"{comp_total:g}h", GREEN)]
    mcells = ""
    for i, (k, v, col) in enumerate(metrics_data):
        bl = "" if i == 0 else f"border-left:1px solid {LINE2};"
        mcells += (f"<td width='1' style='{bl}padding:13px 6px 14px;text-align:center;vertical-align:top'>"
                   f"<div style='font-size:9.5px;font-weight:700;letter-spacing:1px;color:{INK3};"
                   f"text-transform:uppercase'>{k}</div>"
                   f"<div style='font-family:{MONO};font-size:22px;font-weight:700;color:{col};"
                   f"margin-top:6px;line-height:1'>{v}</div></td>")
    kpis = (f"<table role='presentation' width='100%' cellpadding='0' cellspacing='0' "
            f"style='border:1px solid {LINE};border-radius:12px;table-layout:fixed'>"
            f"<tr>{mcells}</tr></table>")

    def _task_html_row(t):
        return (
            f"<tr style='border-top:1px solid {LINE2}'>"
            f"<td style='padding:12px 14px;font-family:{MONO};font-size:13px;"
            f"font-weight:600;white-space:nowrap'>"
            f"<a href='{R._wi_url(app.project, t['id'])}' "
            f"style='color:{VIOLET_INK};text-decoration:none'>{t['id']}</a></td>"
            f"<td style='padding:12px 8px;font-size:13.5px;font-weight:600;color:{INK}'>"
            f"{_html.escape(t['title']) if t['title'] else '—'}</td>"
            f"<td style='padding:12px 8px;font-size:12.5px;color:{INK2}'>"
            f"{_html.escape(t['parent_title']) if t.get('parent_title') else '—'}</td>"
            f"<td style='padding:12px 8px;text-align:center;font-size:11.5px;font-weight:600;"
            f"color:{INK2}'>{_html.escape(str(t['state']))}</td>"
            f"<td style='padding:12px 8px;text-align:right;font-family:{MONO};"
            f"font-size:13.5px;color:{INK}'>{t['original_estimate']:g}</td>"
            f"<td style='padding:12px 14px;text-align:right;font-family:{MONO};"
            f"font-size:13.5px;font-weight:700;color:{INK}'>{t['completed_work']:g}</td></tr>")

    # Capped like every other email report's row list (_plan_html's
    # _STORY_CAP_H, build_report_email's per_story, etc.) — a busy sprint can
    # have a lot of tasks, and listing every one makes the email huge; the
    # attached Excel/PDF is still the full source of truth.
    _TASK_CAP_H = 150
    trows = [_task_html_row(t) for t in tasks[:_TASK_CAP_H]]
    task_tbl = (
        f"<table role='presentation' width='100%' cellpadding='0' cellspacing='0' "
        f"style='border:1px solid {LINE};border-radius:12px;min-width:520px'>"
        f"<tr style='background:{TINT}'>"
        + "".join(f"<td style='padding:10px {p};font-size:10.5px;letter-spacing:.5px;"
                  f"text-transform:uppercase;color:{INK3};font-weight:700;{a}'>{h}</td>"
                  for h, p, a in [("Task", "14px", ""), ("Title", "8px", ""),
                                  ("Story", "8px", ""), ("State", "8px", "text-align:center"),
                                  ("Est.", "8px", "text-align:right"),
                                  ("Completed", "14px", "text-align:right")])
        + "</tr>" + "".join(trows) + "</table>")
    _more_h = (
        f"<div style='margin-top:8px;font-size:11px;color:{INK3}'>&hellip; and "
        f"{len(tasks) - len(trows)} more &middot; see the attached document for the "
        f"full list</div>") if len(tasks) > len(trows) else ""

    scope = (f"{_html.escape(_tm_scope_label(res))} &middot; "
            f"{_html.escape(res.get('assignee_name') or res.get('assignee', ''))}")
    kind = "Task Workload Report"

    # masthead / hero — same layout as the run report / sprint summary / plan emails
    masthead = (
        f"<table role='presentation' width='100%' cellpadding='0' cellspacing='0'><tr>"
        f"<td width='34' valign='middle' style='padding-right:13px'>{E._logo_tag(34)}</td>"
        f"<td valign='middle'>"
        f"<div style='font-size:15px;font-weight:700;color:{INK};letter-spacing:-.2px'>QA Studio</div>"
        f"<div style='font-size:12px;font-weight:700;color:{VIOLET_INK};margin-top:2px'>{kind} &middot; Report</div>"
        f"</td>"
        f"<td valign='middle' align='right' style='font-family:{MONO};font-size:11px;"
        f"color:{INK3};font-weight:700'>{datetime.now():%Y-%m-%d %H:%M}</td></tr></table>")

    hero = (
        f"<span style='display:inline-block;background:{VIOLET_SOFT};color:{VIOLET_INK};"
        f"font-size:11px;font-weight:700;letter-spacing:.4px;padding:5px 12px;"
        f"border-radius:20px'>WORKLOAD SUMMARY</span>"
        f"<div style='font-size:23px;font-weight:700;letter-spacing:-.5px;color:{INK};"
        f"line-height:1.2;margin:14px 0 0'>{kind}</div>"
        f"<div style='font-size:13px;color:{INK2};font-weight:600;margin-top:8px'>{scope}</div>")

    footer = (
        f"<table role='presentation' cellpadding='0' cellspacing='0'><tr>"
        f"<td valign='middle' style='padding-right:9px'>{E._logo_tag(20)}</td>"
        f"<td valign='middle' style='font-size:11.5px;font-weight:600;color:{INK3}'>"
        f"Generated by QA Studio &middot; Azure DevOps + AI</td></tr></table>"
        f"<div style='font-family:{MONO};font-size:11px;color:{INK3};margin-top:8px;"
        f"line-height:1.6'>The full report is attached as Excel and PDF.</div>")

    return f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
</head>
<body style='margin:0;padding:0;background:{PAPER};-webkit-text-size-adjust:100%'>
<center style='width:100%;background:{PAPER}'>
<table role='presentation' width='100%' cellpadding='0' cellspacing='0' style='background:{PAPER}'><tr>
<td align='center' style='padding:26px 12px 48px'>
<table role='presentation' width='680' cellpadding='0' cellspacing='0' style='width:680px;max-width:680px;background:{CARD};border:1px solid #DEDDE6;border-radius:16px;overflow:hidden;font-family:{UI};color:{INK}'>
  <tr><td style='height:3px;line-height:3px;font-size:0;background:{VIOLET}'>&nbsp;</td></tr>
  <tr><td style='padding:24px 32px 0'>{masthead}</td></tr>
  <tr><td style='padding:18px 32px 4px'>{hero}</td></tr>
  <tr><td style='padding:18px 32px 0'>{kpis}</td></tr>
  <tr><td style='padding:26px 32px 6px;border-top:1px solid {LINE}'>
    {_sec_head(VIOLET, 'Tasks', len(tasks))}
    <div style='margin-top:14px;overflow-x:auto;-webkit-overflow-scrolling:touch'>{task_tbl}</div>
    {_more_h}
    </td></tr>
  <tr><td style='padding:20px 32px 26px;border-top:1px solid {LINE};background:{TINT}'>{footer}</td></tr>
</table>
</td></tr></table></center></body></html>"""


def _tm_export_row(app, res):
    """Excel / PDF / JSON export buttons for the workload report — same
    save-dialog + status-notify pattern as regression.py's _export_row, minus
    Word (kept out to control scope; can be added the same way if wanted)."""
    import regression as R

    def _notify(kind, text):
        app._tm_report_export_msg = (kind, text)
        app.ui_safe(lambda k=kind, t=text: (app._toast(t) if k == "ok" else app._err(t)))

    _fns = {"xlsx": _tm_export_xlsx, "pdf": _tm_export_pdf, "json": _tm_export_json}

    def _go(fmt):
        def _do(e):
            def work():
                try:
                    dest = R._ask_save_path(fmt, _tm_stamp(app, res) + "." + fmt)
                    if dest is None:
                        return
                    path = _fns[fmt](app, res)
                    if dest and dest is not False:
                        if not dest.lower().endswith("." + fmt):
                            dest += "." + fmt
                        import shutil
                        if os.path.abspath(dest) != os.path.abspath(path):
                            shutil.move(path, dest)
                        path = dest
                    try:
                        import platform_caps as _pc
                        if _pc.is_mobile():
                            _pc.reveal_export(app.page, path)
                            _notify("ok", f"{fmt.upper()} ready — choose where to save or send it.")
                        else:
                            _pc.open_folder(os.path.dirname(path))
                            _notify("ok", f"Saved: {path}")
                    except Exception:
                        _notify("ok", f"Saved: {path}")
                except ModuleNotFoundError as md:
                    _notify("err", f"Missing dependency: {md.name}")
                except Exception as ex:
                    _notify("err", f"Export failed: {str(ex)[:160]}")
            threading.Thread(target=work, daemon=True).start()
        return _do

    def _btn(label, icon, color, fmt):
        return ft.OutlinedButton(
            content=ft.Row([ft.Icon(icon, size=17, color=color),
                           ft.Text(label, size=13.5, weight=ft.FontWeight.W_600, color=T.INK)],
                          spacing=8, tight=True),
            on_click=_go(fmt), height=44,
            style=ft.ButtonStyle(bgcolor={"": T.CARD}, side=ft.BorderSide(1, T.BORDER),
                                 shape=ft.RoundedRectangleBorder(radius=T.R),
                                 padding=ft.Padding.symmetric(horizontal=15, vertical=0)))

    return ft.Row([
        _btn("Excel", ft.Icons.TABLE_CHART, T.GREEN, "xlsx"),
        _btn("PDF", ft.Icons.PICTURE_AS_PDF, T.RED, "pdf"),
        _btn("JSON", ft.Icons.DATA_OBJECT, T.STORY, "json"),
    ], spacing=8, wrap=True)


def _tm_email(app, res):
    if getattr(app, "_tm_report_emailing", False):
        app._toast("Already sending — please wait…")
        return
    to = [a.strip() for a in re.split(r"[,\s;]+", (app._tm_report_email_to or "")) if a.strip()]
    if not to:
        app._err("Enter at least one recipient email.")
        return
    if not getattr(E, "GMAIL_APP_PASS", ""):
        app._err("Set the Gmail App Password on the Setup screen first.")
        return
    app._tm_report_email_to = ", ".join(to)
    app._tm_report_emailing = True
    app._toast("Sending the task report…")
    _b = getattr(app, "_tm_report_send_btn", None)
    if _b is not None:
        try:
            _b.disabled = True; _b.opacity = 0.55; _b.update()
        except Exception:
            pass

    def work():
        try:
            attach = []
            for fn in (_tm_export_xlsx, _tm_export_pdf):
                try:
                    attach.append(fn(app, res))
                except Exception:
                    pass
            subj = (f"Task Workload Report — {res.get('assignee_name') or res.get('assignee', '')} — "
                   f"{_tm_scope_label(res)}")
            ok, err = E.send_report(to, subj, _tm_report_html(app, res), attachments=attach)
            kind = "ok" if ok else "err"
            text = f"Emailed to {', '.join(to)}" if ok else (err or "Email failed.")
        except Exception as ex:
            kind, text = "err", f"Email failed: {str(ex)[:160]}"
        app._tm_report_emailing = False

        def _fin(k=kind, t=text):
            _b2 = getattr(app, "_tm_report_send_btn", None)
            if _b2 is not None:
                try:
                    _b2.disabled = False; _b2.opacity = 1.0; _b2.update()
                except Exception:
                    pass
            app._toast(t) if k == "ok" else app._err(t)
        app.ui_safe(_fin)
    threading.Thread(target=work, daemon=True).start()


def screen(app):
    _init(app)
    import regression as R
    from main import (card, sec_head, field_label, primary_btn, green_btn, hover_field,
                      searchable_dropdown)

    if not app.readonly and not (app.connected and app.project):
        return R.locked_state(
            app, "Task Manager",
            "Per-user task workload reports and bulk child-task creation",
            "Connect your Azure DevOps account on the Setup screen first.",
            icon=ft.Icons.FACT_CHECK_OUTLINED)

    # Bumped on every FULL rebuild of this screen — captured below as
    # `_my_epoch` by _sync_report_dynamic/_sync_ct_dynamic. The real bug this
    # guards against: _load_iterations/_load_members/_load_stories (right
    # below) and _set_iteration (switching sprint) all trigger a full
    # app.render() independently, on their own background thread's own
    # timeline — completely unrelated to whatever Calculate is doing. If one
    # of those fires WHILE a Calculate fetch from THIS screen() call is still
    # in flight (very likely right after switching sprints, since that
    # restarts _load_stories, which is exactly what a user is likely to do
    # right before clicking Calculate again), the fetch's completion later
    # calls ITS on_update closure (_sync_report_dynamic), which was captured
    # back when THIS screen() ran — but by then a NEWER screen() call has
    # replaced the whole page, so that closure's cell is a dead, detached
    # Container. cell.update() on it silently no-ops (caught by the bare
    # except in the sync functions below), so the fetch's busy=False/result
    # never reaches the CURRENTLY VISIBLE cell — which is stuck showing
    # whatever it looked like at that last real render (a disabled
    # "Calculating…" button), forever, with no exception ever logged
    # anywhere. Checking the epoch lets a stale closure detect this and fall
    # back to a fresh app.render() instead of touching the dead cell.
    app._tm_screen_epoch = getattr(app, "_tm_screen_epoch", 0) + 1
    _my_epoch = app._tm_screen_epoch

    # A true Viewer (readonly) gets a genuinely inert screen, not just a
    # blocked Calculate/Create-tasks click at the end — every interactive
    # input (sprint/date pickers, the scope toggle, both "assign to"
    # dropdowns, the story picker, and every per-story task field) is
    # disabled too, so there's nothing to click that LOOKS editable but
    # silently goes nowhere. `_calc_report`/`_create_tasks` still check this
    # themselves as a second line of defense (in case a click reaches them
    # some other way), but the fields shouldn't invite the click in the
    # first place.
    _ro = getattr(app, "readonly", False)

    _load_iterations(app)
    _load_members(app)
    _load_stories(app)

    members = getattr(app, "_members_cache", None) or []

    def _member_opts():
        # A FRESH list of DropdownOption instances every call — these two
        # dropdowns (report_user_dd, ct_user_dd) must never share the same
        # control objects. A single shared list rendered fine with plain
        # ft.Dropdown, but silently broke Section 2 once ct_user_dd switched
        # to searchable_dropdown (editable=True/enable_filter=True): Flet
        # parents each option control once, so re-parenting the same
        # instances under a second Dropdown left that second control (and
        # everything built after it) failing to render — the exact grey/
        # blank "Create child tasks" card seen live. Rebuilding the options
        # per-dropdown is the fix.
        return [ft.DropdownOption(key=m["email"], text=m["name"]) for m in members]

    iter_options = [ft.DropdownOption(key=it["path"],
                                      text=(_sprint_num(it["name"]) or it["name"]))
                    for it in app._tm_iterations]

    # ── Section 1: Sprint/Date-range scope + workload report, merged into one
    # card. Previously two separate cards (a "Sprint" picker card and a
    # "Task workload report" card below it) — merged per the user's request,
    # since they're really one decision ("what period, who, then Calculate"),
    # not two. Rebuilt in place via _report_dynamic_cell/_sync_report_dynamic
    # (same technique as Section 2's _ct_dynamic_cell below) instead of a
    # full app.render() on every toggle flip / date pick / Assigned-to pick /
    # Calculate click — this was the single biggest source of whole-screen
    # re-renders on this screen.
    def _set_iteration(e):
        app._tm_iteration = e.control.value or ""
        app._tm_ct_selected = []
        app._tm_ct_rows = {}
        app._tm_report_result = None
        app._tm_report_msg = None
        app._tm_ct_result = None
        app._tm_ct_msg = None
        # A sprint change reshapes THREE areas at once (report reset, story
        # list reload, bulk-create reset) and Section 2's stories need a
        # fresh _load_stories() call, which only happens on screen()'s own
        # top-level path — a full render is genuinely needed here, unlike
        # the smaller per-field updates below.
        app.render()

    # ── Sprint vs. Date range — an explicit toggle, not "fill in whichever".
    # Calculate is scoped by exactly one of these at a time, picked via
    # _tm_scope_mode ("sprint"|"dates"); the inactive side keeps its state
    # (switching back and forth never loses what was typed/picked into it),
    # it's just fully INTERACTIVE-DISABLED — not merely dimmed — while it's
    # not the active mode, at the user's explicit request. This does mean
    # Section 2's bulk task creation (which has no date-range equivalent and
    # always needs a sprint) can only be pointed at whatever sprint was last
    # picked while in "Sprint" mode: if that field is currently disabled
    # (i.e. "Date range" is active) and no sprint has ever been picked,
    # Section 2 has nothing to load stories from until the toggle is flipped
    # back to "Sprint" and one is chosen. Accepted trade-off — flipping the
    # toggle never clears the underlying value, so this is a one-time
    # "switch back to pick a sprint" step, not a dead end.
    def _set_scope_mode(mode):
        def _h(e):
            app._tm_scope_mode = mode
            _sync_report_dynamic()
            # Section 2's locked/unlocked state depends on this same toggle
            # (see _build_ct_section below) — without this call it only ever
            # picks up the flip the next time something ELSE happens to force
            # a full app.render() (e.g. switching sprints), which from the
            # toggle alone looked like clicking it "did nothing" to Section 2.
            _sync_ct_section()
        return _h

    def _set_start_date(v):
        app._tm_start_date = v

    def _set_end_date(v):
        app._tm_end_date = v

    def _clear_dates(e):
        # The date fields (via _date_field's ft.DatePicker) can only be SET,
        # never unset, by picking a day — there was no way to get back to
        # "no date range" once one was touched. Gives explicit control back
        # for anyone who wants the fields visibly empty again.
        app._tm_start_date = ""
        app._tm_end_date = ""
        _sync_dates()

    def _sync_dates():
        _sync_report_dynamic()

    def _build_dates_row(disabled):
        start = (app._tm_start_date or "").strip()
        end = (app._tm_end_date or "").strip()
        has_dates = bool(start or end)
        invalid_range = bool(start and end and end < start)

        # Clear now sits INLINE with the two date boxes (bottom-aligned to
        # them via vertical_alignment=END) instead of living in a header row
        # above them — reads as "part of the input", not a separate action
        # bar. Icon-only (a tooltip covers the label) so it doesn't crowd
        # two 150px-wide boxes.
        clear_btn = ft.IconButton(
            icon=ft.Icons.CLOSE_ROUNDED, icon_size=16, icon_color=T.INK_3,
            tooltip="Clear both dates", on_click=_clear_dates,
            width=40, height=40,
            style=ft.ButtonStyle(
                shape=ft.RoundedRectangleBorder(radius=T.R),
                bgcolor={"": T.CARD_2}, side=ft.BorderSide(1, T.BORDER)))

        _sd_field = _date_field(app, "Start date", app._tm_start_date, _set_start_date,
                                on_after=_sync_dates, disabled=disabled)
        _ed_field = _date_field(app, "End date", app._tm_end_date, _set_end_date,
                                on_after=_sync_dates, disabled=disabled, min_date=(start or None))
        import platform_caps as _pc_tm
        if _pc_tm.is_mobile():
            # Each date box is a fixed 150px (_date_field's own width) — two
            # of them plus the 40px clear button (150+150+40+spacing ≈ 380px)
            # left no margin at all against a ~390px phone once this card's
            # own padding is subtracted, so End date (last in the Row, no
            # wrap) rendered fully or partly off-screen — reported live as
            # "end date not displays". Wrap instead of a single unwrapping
            # Row so End date (and the clear button) drop to a second line
            # rather than being clipped.
            fields_row = ft.Row(
                [_sd_field, _ed_field] + ([clear_btn] if (has_dates and not disabled) else []),
                spacing=10, run_spacing=10, wrap=True,
                vertical_alignment=ft.CrossAxisAlignment.END)
        else:
            fields_row = ft.Row(
                [_sd_field, _ed_field] + ([clear_btn] if (has_dates and not disabled) else []),
                spacing=10, vertical_alignment=ft.CrossAxisAlignment.END)

        parts = [
            ft.Text("Date range", size=12, weight=ft.FontWeight.BOLD,
                   color=(T.INK_3 if disabled else T.INK_2)),
            ft.Container(height=8),
            fields_row,
        ]
        if invalid_range and not disabled:
            parts.append(ft.Container(height=6))
            parts.append(ft.Row([
                ft.Icon(ft.Icons.ERROR_OUTLINE, size=13, color=T.RED),
                ft.Text("End date must be on or after the start date.",
                       size=11, color=T.RED),
            ], spacing=6))
        return ft.Column(parts, spacing=0)

    def _set_report_user(e):
        app._tm_report_user = e.control.value or ""
        _sync_report_dynamic()

    _report_dynamic_cell = [None]

    def _build_report_dynamic():
        mode = getattr(app, "_tm_scope_mode", "sprint")
        sprint_active = (mode != "dates")

        # Setup screen's segmented-toggle look (main.py's _tool_segment /
        # _lang_segment) — a soft-violet fill + violet border + faint glow
        # on the selected side, not a solid-fill pill — reused here so this
        # toggle matches the one visual language every other multi-option
        # switch in the app already uses.
        def _scope_seg(label, icon, key):
            active = (mode == key)
            c = ft.Container(
                ft.Row([ft.Icon(icon, size=15,
                               color=(T.VIOLET_INK if active else T.INK_2)),
                       ft.Text(label, size=12.5, weight=ft.FontWeight.BOLD,
                              color=(T.VIOLET_INK if active else T.INK_2))],
                      spacing=7, alignment=ft.MainAxisAlignment.CENTER, tight=True),
                expand=True, height=40, alignment=ft.Alignment.CENTER,
                padding=ft.Padding.symmetric(horizontal=9),
                bgcolor=(T.VIOLET_SOFT if active else None), border_radius=T.R_SM,
                border=ft.Border.all(1, T.VIOLET if active else ft.Colors.TRANSPARENT),
                animate=140,
                shadow=(ft.BoxShadow(blur_radius=10, spread_radius=-4,
                                     color=ft.Colors.with_opacity(0.35, T.VIOLET),
                                     offset=ft.Offset(0, 3)) if active else None),
                opacity=(0.55 if _ro else 1.0),
                on_click=(None if _ro else _set_scope_mode(key)))
            if not active and not _ro:
                def _h(e, _c=c):
                    try:
                        _c.bgcolor = (ft.Colors.with_opacity(0.55, T.CARD)
                                      if e.data in (True, "true", "True") else None)
                        _c.update()
                    except Exception:
                        pass
                c.on_hover = _h
            return c

        scope_toggle = ft.Container(
            ft.Row([_scope_seg("Sprint", ft.Icons.VIEW_WEEK_OUTLINED, "sprint"),
                   _scope_seg("Date range", ft.Icons.DATE_RANGE_OUTLINED, "dates")],
                  spacing=4),
            padding=4, bgcolor=T.CARD_2, border_radius=T.R,
            border=ft.Border.all(1, T.BORDER_2))

        # Built FRESH every rebuild (own options list, own instance) — same
        # rule as report_user_dd/ct_user_dd elsewhere in this file: never
        # reuse a control instance across a cell's rebuilds. `disabled` is
        # now unconditional on the toggle's mode (not just dimmed) — picking
        # a sprint is simply not possible while "Date range" is active.
        iter_dd = ft.Dropdown(
            value=app._tm_iteration or None, options=iter_options,
            hint_text=("Loading sprints…" if app._tm_iter_loading else "Select sprint…"),
            text_size=13, border_color=T.BORDER, focused_border_color=T.VIOLET,
            border_radius=T.R, content_padding=ft.Padding.symmetric(vertical=8, horizontal=10),
            on_select=_set_iteration,
            disabled=(app._tm_iter_loading or not sprint_active or _ro))

        scope_row = ft.Row([
            ft.Column([field_label("Sprint", req=True), hover_field(iter_dd)],
                     spacing=6, expand=True,
                     opacity=(1.0 if (sprint_active and not _ro) else 0.45)),
            ft.Column([_build_dates_row(disabled=(sprint_active or _ro))],
                     spacing=0, expand=True),
        ], spacing=20, vertical_alignment=ft.CrossAxisAlignment.START)

        # Built FRESH every rebuild (own option list via _member_opts(), own
        # control instance) — never reused across rebuilds. Reusing the same
        # Dropdown/option instances across repeated re-parenting is exactly
        # what silently broke Section 2 earlier this session; building fresh
        # every time is the proven-safe pattern here.
        report_user_dd = searchable_dropdown(
            value=app._tm_report_user or None, options=_member_opts(),
            hint_text="Type to search a person…", text_size=13, border_color=T.BORDER,
            focused_border_color=T.VIOLET, border_radius=T.R,
            content_padding=ft.Padding.symmetric(vertical=8, horizontal=10),
            on_select=_set_report_user, disabled=_ro)

        it = app._tm_iteration
        start = (app._tm_start_date or "").strip()
        end = (app._tm_end_date or "").strip()
        use_dates = (mode == "dates")
        scope_ok = (bool(it) if not use_dates
                   else bool(start and end and end >= start))
        _can_calc = (scope_ok and bool(app._tm_report_user)
                    and not app._tm_report_busy and not _ro)
        calc_btn = primary_btn(
            "Calculating…" if app._tm_report_busy else "Calculate",
            icon=ft.Icons.CALCULATE_OUTLINED,
            on_click=((lambda e: _calc_report(app, on_update=_sync_report_dynamic))
                     if _can_calc else None))
        try:
            calc_btn.opacity = 1.0 if _can_calc else 0.45
        except Exception:
            pass

        parts = [card(ft.Column([
            sec_head("1", "Task workload report"),
            ft.Container(height=14),
            scope_toggle,
            ft.Container(height=16),
            scope_row,
            ft.Container(height=1, bgcolor=T.BORDER_2,
                        margin=ft.Margin.symmetric(vertical=18)),
            ft.Column([field_label("Assigned to", req=True), hover_field(report_user_dd)],
                      spacing=6),
            ft.Container(height=14),
            calc_btn,
        ], spacing=0))]

        banner = _status_banner(app._tm_report_msg, app._tm_report_busy)
        if banner:
            parts.append(banner)

        res = app._tm_report_result
        if res and res.get("count", 0) > 0:
            def _recalc_totals():
                # Recomputed from the LOCAL task list, not re-fetched from
                # Azure — mirrors the Sprint Summary modal's own inline
                # _delete_story (modals.py): every downstream reader of `res`
                # (the KPI tiles below, the 170h/month benchmark, and every
                # export/email builder — they all just call res.get(...) at
                # click-time) sees the updated numbers for free once these
                # three keys are mutated in place, no separate "recalc" logic
                # needed anywhere else.
                ts = res["tasks"]
                res["count"] = len(ts)
                res["total_original_estimate"] = sum(t.get("original_estimate", 0) for t in ts)
                res["total_completed_work"] = sum(t.get("completed_work", 0) for t in ts)

            def _delete_task(tid):
                def _do():
                    res["tasks"] = [t for t in res["tasks"] if t["id"] != tid]
                    _recalc_totals()
                    app.render()
                    try:
                        app._toast(f"Removed task {tid} from this report.")
                    except Exception:
                        pass
                def _d(e):
                    if _ro:
                        return app._toast("Read-only — your role can't modify the report.")
                    app._confirm(
                        "Remove task?",
                        f"Remove task {tid} from this report and recalculate the "
                        "totals? This doesn't change anything in Azure DevOps.",
                        _do, yes_label="Remove")
                return _d

            # TASK column: expand=True on desktop (fills the wide window), a
            # FIXED width on mobile so it doesn't collapse to 0 inside the
            # horizontal-scroll wrapper below (which is what char-wrapped the
            # "TASK" header to "T/A/S/K"). Header uses the same width.
            import platform_caps as _pc_tm
            _task_w = 220 if _pc_tm.is_mobile() else None
            rows = [ft.Container(
                ft.Row([
                    R._id_link(app, t["id"], tooltip=f"Open task {t['id']} in Azure DevOps",
                              color=T.VIOLET_INK, weight=ft.FontWeight.BOLD,
                              width=70, font_family=T.F_MONO, size=12.5),
                    ft.Text(t["title"] or "—", size=12.5, color=T.INK,
                           expand=(None if _task_w else True), width=_task_w,
                           max_lines=1, overflow=ft.TextOverflow.ELLIPSIS),
                    ft.Text(t.get("parent_title") or "—", size=12, color=T.INK_3, width=180,
                           max_lines=1, overflow=ft.TextOverflow.ELLIPSIS),
                    ft.Text(t["state"], size=12, color=T.INK_2, width=90),
                    ft.Text(f"{t['original_estimate']:g}", size=12.5, color=T.INK, width=90,
                           text_align=ft.TextAlign.RIGHT),
                    ft.Text(f"{t['completed_work']:g}", size=12.5, color=T.INK, width=90,
                           text_align=ft.TextAlign.RIGHT),
                    ft.IconButton(
                        icon=ft.Icons.DELETE_OUTLINE, icon_size=18, icon_color=T.RED,
                        tooltip="Remove from this report", disabled=_ro,
                        on_click=_delete_task(t["id"]),
                        width=34, height=34,
                        style=ft.ButtonStyle(padding=ft.Padding.all(0),
                                             shape=ft.RoundedRectangleBorder(radius=8))),
                ], spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                padding=ft.Padding.symmetric(vertical=9, horizontal=12),
                bgcolor=(T.CARD if i % 2 == 0 else T.CARD_2),
                border=ft.Border.only(bottom=ft.BorderSide(1, T.BORDER_2)))
                for i, t in enumerate(res["tasks"])]
            hdr = ft.Container(
                ft.Row([
                    ft.Text("ID", size=10.5, weight=ft.FontWeight.BOLD, color=T.INK_2, width=70),
                    ft.Text("TASK", size=10.5, weight=ft.FontWeight.BOLD, color=T.INK_2,
                           expand=(None if _task_w else True), width=_task_w),
                    ft.Text("STORY", size=10.5, weight=ft.FontWeight.BOLD, color=T.INK_2, width=180),
                    ft.Text("STATE", size=10.5, weight=ft.FontWeight.BOLD, color=T.INK_2, width=90),
                    ft.Text("ORIG. EST.", size=10.5, weight=ft.FontWeight.BOLD, color=T.INK_2,
                           width=90, text_align=ft.TextAlign.RIGHT),
                    ft.Text("COMPLETED", size=10.5, weight=ft.FontWeight.BOLD, color=T.INK_2,
                           width=90, text_align=ft.TextAlign.RIGHT),
                    ft.Container(width=34),
                ], spacing=10),
                padding=ft.Padding.symmetric(vertical=11, horizontal=12),
                bgcolor=T.CARD_2, border=ft.Border.only(bottom=ft.BorderSide(1, T.BORDER)))
            est_total = res["total_original_estimate"]
            comp_total = res["total_completed_work"]
            pct = (comp_total / est_total) if est_total > 0 else (1.0 if comp_total > 0 else 0.0)
            # Completed Work can genuinely run past Original Estimate — people
            # log more actual hours than they first estimated all the time —
            # so `pct` itself is left uncapped (it still drives the GREEN
            # "done" color via `pct >= 1` below). But the DISPLAYED percentage
            # is capped at 100: a progress readout showing "102% complete"
            # reads as a bug even when the ratio behind it is real. The
            # overrun itself isn't hidden, just shown separately below.
            pct_disp = min(100, round(pct * 100))
            over_by = max(0.0, comp_total - est_total)
            remaining = max(0.0, est_total - comp_total)

            # 170h/month benchmark — a SEPARATE "remaining" figure from the
            # one above (never replaces it): how many hours are left before
            # this person hits a normal month's workload, prorated over
            # whichever period THIS result actually covers. Uses the period
            # the result was calculated for, not necessarily whatever the
            # toggle/dropdown/date fields show live right now (those could
            # have moved on since — same trust-the-result convention est_total/
            # comp_total above already follow).
            #
            # Date range mode ONLY. Sprints almost never line up with a
            # calendar month (two-week sprints are the norm, and even
            # "monthly" sprints rarely run exactly 30.44 days), so prorating
            # a 170h/MONTH benchmark over a sprint's own start/finish dates
            # produces a number that LOOKS precise but isn't a meaningful
            # comparison — it would never actually be accurate. The tile
            # (and its caption) is simply omitted in Sprint mode.
            if res.get("iteration"):
                bench_start = bench_end = None
            else:
                bench_start, bench_end = app._tm_start_date, app._tm_end_date
            bench_days = _tm_period_days(bench_start, bench_end)
            bench_hours = _tm_benchmark_hours(bench_days)
            # Rounded to whole hours for display — a benchmark this approximate
            # (170h/month is itself a round, average figure) showing 4+ decimal
            # places ("52.6189 h") read as false precision, not accuracy.
            bench_remaining = (round(max(0.0, bench_hours - comp_total))
                              if bench_hours is not None else None)
            bench_period_label = (
                f"{_tm_fmt_date(bench_start)} – {_tm_fmt_date(bench_end)} "
                f"· {bench_days} day{'s' if bench_days != 1 else ''}"
            ) if bench_days else None

            # KPI strip — same tile/gradient/shadow language as Sprint Plan's
            # _kpis() strip. IMPORTANT: no wrap=True here (unlike this file's
            # export-button Row, where it's safe) — _kpi_tile sets
            # expand=True on its Container, which Flet compiles to a Flutter
            # Expanded, and Expanded REQUIRES a Flex (plain Row/Column)
            # ancestor, not a Wrap. This exact combination crashed the client
            # earlier this session ("WrapParentData is not a subtype of
            # FlexParentData"); keep this Row wrap-free.
            _kpi_tiles = [
                R._kpi_tile("TASKS", str(res["count"])),
                R._kpi_tile("ORIGINAL ESTIMATE", f"{est_total:g} h"),
                R._kpi_tile("COMPLETED WORK", f"{comp_total:g} h", T.GREEN),
                R._kpi_tile("REMAINING", f"{remaining:g} h"),
            ]
            if bench_remaining is not None:
                _kpi_tiles.append(R._kpi_tile("REMAINING (170H/MO)", f"{bench_remaining:g} h"))
            # Responsive: equal-share on desktop; fixed-width wrapping tiles on
            # a phone (the plain expand Row char-wrapped every label). See
            # R.kpi_row.
            kpi_strip = R.kpi_row(_kpi_tiles, spacing=14)
            # Small caption naming exactly what period the 170h/month tile
            # above was prorated over — without this, "REMAINING (170H/MO)"
            # alone doesn't say whether that's this sprint, this week, or
            # this month, especially since it can differ from the sprint's
            # OWN dates shown nowhere else on this card.
            bench_caption = (
                ft.Container(
                    ft.Row([
                        ft.Icon(ft.Icons.CALENDAR_MONTH_OUTLINED, size=12.5, color=T.INK_3),
                        ft.Text(f"170h/month benchmark calculated for {bench_period_label}",
                               size=11, color=T.INK_3, weight=ft.FontWeight.W_600),
                    ], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    margin=ft.Margin.only(top=10))
                if bench_period_label else ft.Container(height=0))
            progress = ft.Container(
                ft.Column([
                    ft.Row([
                        ft.Text("PROGRESS", size=10.5, weight=ft.FontWeight.BOLD, color=T.INK_3,
                               expand=True),
                        ft.Text(
                            f"{pct_disp}% complete" + (f"  ·  {over_by:g}h over estimate"
                                                       if over_by > 0 else ""),
                            size=12, weight=ft.FontWeight.BOLD,
                            color=(T.GREEN if pct >= 1 else T.VIOLET_INK)),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    ft.Container(height=8),
                    R._bar(min(1.0, pct), (T.GREEN if pct >= 1 else T.VIOLET), 8),
                ], spacing=0),
                padding=14, bgcolor=T.CARD, border=ft.Border.all(1, T.BORDER_2),
                border_radius=T.R, margin=ft.Margin.only(top=14))
            parts.append(ft.Container(ft.Column([kpi_strip, bench_caption, progress], spacing=0),
                                      margin=ft.Margin.only(top=14)))
            # Column widths: 70 + TASK(220) + 180 + 90 + 90 + 90 + 34, plus
            # 6×10 spacing + 2×12 padding ≈ 838. On mobile the whole table
            # scrolls sideways at that width so no column is crushed.
            _rows_col = ft.Column(rows, spacing=0, scroll=ft.ScrollMode.AUTO,
                                  height=min(360, 48 * len(rows) + 20))
            parts.append(ft.Container(
                R.hscroll_table(hdr, _rows_col, 838),
                margin=ft.Margin.only(top=14), border=ft.Border.all(1, T.BORDER),
                border_radius=T.R, clip_behavior=ft.ClipBehavior.ANTI_ALIAS))

            # Export (Excel / PDF / JSON) + email — same pattern as Sprint Plan.
            parts.append(ft.Container(
                ft.Column([
                    ft.Text("EXPORT", size=10.5, weight=ft.FontWeight.BOLD, color=T.INK_3),
                    ft.Container(height=8),
                    _tm_export_row(app, res),
                ], spacing=0), margin=ft.Margin.only(top=16)))

            def _email(e, _res=res):
                _tm_email(app, _res)

            app._tm_report_send_btn = green_btn("Email report", icon=ft.Icons.SEND,
                                                on_click=_email)
            email_picker = R.email_recipient_picker(
                app, "_tm_report_email_to", is_open_key="_tm_report_email_open",
                sync_key="tm_report_emails", trailing=app._tm_report_send_btn)
            parts.append(ft.Container(
                ft.Column([
                    ft.Text("EMAIL", size=10.5, weight=ft.FontWeight.BOLD, color=T.INK_3),
                    ft.Container(height=8),
                    email_picker,
                ], spacing=0), margin=ft.Margin.only(top=16)))

        return ft.Column(parts, spacing=0)

    def _sync_report_dynamic():
        if getattr(app, "_tm_screen_epoch", 0) != _my_epoch:
            # A newer screen() call has already replaced the page since this
            # closure was captured (e.g. _load_stories finished mid-Calculate
            # right after switching sprints) — _report_dynamic_cell[0] here
            # is a dead Container no longer attached to the live page.
            # Updating it would silently no-op below; only a fresh render
            # can actually show the current state now.
            app.render()
            return
        cell = _report_dynamic_cell[0]
        if cell is None:
            return
        try:
            cell.content = _build_report_dynamic()
            cell.update()
        except Exception:
            pass

    report_dynamic = ft.Container(_build_report_dynamic())
    _report_dynamic_cell[0] = report_dynamic
    report_body = [report_dynamic]

    # ── Section 2: bulk child-task creation ────────────────────────────────
    _ct_dynamic_cell = [None]   # rebuilt-in-place: title rows + Create button

    def _build_ct_dynamic():
        stories_by_id = {s["id"]: s for s in app._tm_ct_stories}
        story_blocks = []
        for sid in app._tm_ct_selected:
            st = stories_by_id.get(sid, {})
            rows = app._tm_ct_rows.setdefault(sid, [])
            if not rows:
                rows.append(_tm_default_row(st))

            row_blocks = []
            for idx, row in enumerate(rows):
                tf = ft.TextField(
                    value=row.get("title", ""), on_change=_set_row_field(sid, idx, "title"),
                    text_size=12.5, dense=True, border_color=T.BORDER,
                    focused_border_color=T.VIOLET, border_radius=T.R, disabled=_ro,
                    content_padding=ft.Padding.symmetric(vertical=8, horizontal=10))
                # Only the FIRST row for a story shows the clickable story-id
                # link (it's the same story for every row in the batch); the
                # rest just show a small "#2", "#3"... index so the batch
                # still reads as a numbered list without repeating the link.
                lead_cell = (
                    R._id_link(app, sid, color=T.VIOLET_INK, weight=ft.FontWeight.BOLD,
                              width=70, font_family=T.F_MONO, size=12)
                    if idx == 0 else
                    ft.Text(f"#{idx + 1}", size=11.5, color=T.INK_3, width=70,
                           font_family=T.F_MONO, text_align=ft.TextAlign.CENTER))
                remove_btn = (
                    ft.IconButton(icon=ft.Icons.CLOSE_ROUNDED, icon_size=15,
                                 icon_color=T.INK_3, tooltip="Remove this task",
                                 on_click=(None if _ro else _remove_row(sid, idx)),
                                 disabled=_ro, width=30, height=30)
                    if len(rows) > 1 else ft.Container(width=30))
                title_row = ft.Row([
                    lead_cell,
                    ft.Container(hover_field(tf), expand=True),
                    remove_btn,
                ], spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER)

                due_field = _date_field(app, "Due date", row.get("due", ""),
                                        _set_row_field(sid, idx, "due"),
                                        on_after=_sync_ct_dynamic, disabled=_ro)
                est_tf = ft.TextField(
                    value=row.get("estimate", ""), on_change=_set_row_field(sid, idx, "estimate"),
                    text_size=12.5, dense=True, border_color=T.BORDER,
                    focused_border_color=T.VIOLET, border_radius=T.R, width=110, disabled=_ro,
                    hint_text="hours", keyboard_type=ft.KeyboardType.NUMBER,
                    content_padding=ft.Padding.symmetric(vertical=8, horizontal=10))
                comp_tf = ft.TextField(
                    value=row.get("completed", ""), on_change=_set_row_field(sid, idx, "completed"),
                    text_size=12.5, dense=True, border_color=T.BORDER,
                    focused_border_color=T.VIOLET, border_radius=T.R, width=110, disabled=_ro,
                    hint_text="hours", keyboard_type=ft.KeyboardType.NUMBER,
                    content_padding=ft.Padding.symmetric(vertical=8, horizontal=10))
                est_block = ft.Column([ft.Text("Original estimate", size=10.5,
                                              weight=ft.FontWeight.BOLD, color=T.INK_3),
                                       hover_field(est_tf)], spacing=6, tight=True)
                comp_block = ft.Column([ft.Text("Completed work", size=10.5,
                                                weight=ft.FontWeight.BOLD, color=T.INK_3),
                                        hover_field(comp_tf)], spacing=6, tight=True)
                import platform_caps as _pc_tm
                if _pc_tm.is_mobile():
                    # Desktop packs due_field(150) + a leading 80px alignment spacer
                    # + two 110-wide estimate boxes into one Row — 492px minimum
                    # before spacing, well past a ~390px phone, which is why Due
                    # date/Original estimate were getting clipped. Mobile drops the
                    # purely-cosmetic alignment spacer and wraps to two rows: Due
                    # date alone, then the two 110px estimate boxes together
                    # (220px + spacing comfortably fits).
                    fields_row = ft.Column([
                        due_field,
                        ft.Container(height=10),
                        ft.Row([est_block, comp_block], spacing=14,
                               vertical_alignment=ft.CrossAxisAlignment.END),
                    ], spacing=0)
                else:
                    fields_row = ft.Row([
                        ft.Container(width=80),
                        due_field, est_block, comp_block,
                    ], spacing=14, vertical_alignment=ft.CrossAxisAlignment.END)

                row_blocks.append(ft.Container(
                    ft.Column([title_row, ft.Container(height=10), fields_row], spacing=0),
                    padding=ft.Padding.only(top=(10 if idx else 0), bottom=10),
                    border=(ft.Border.only(bottom=ft.BorderSide(1, T.BORDER_2))
                           if idx < len(rows) - 1 else None)))

            at_max = len(rows) >= _TM_CT_MAX_ROWS
            add_btn = ft.TextButton(
                content=ft.Row([
                    ft.Icon(ft.Icons.ADD_CIRCLE_OUTLINE, size=16,
                           color=(T.INK_3 if (at_max or _ro) else T.VIOLET_INK)),
                    ft.Text(("Add another task" if not at_max
                            else f"Max {_TM_CT_MAX_ROWS} tasks per story"),
                           size=12, weight=ft.FontWeight.W_600,
                           color=(T.INK_3 if (at_max or _ro) else T.VIOLET_INK)),
                ], spacing=6, tight=True),
                on_click=(None if (at_max or _ro) else _add_row(sid)), disabled=(at_max or _ro))

            story_blocks.append(ft.Container(
                ft.Column([
                    ft.Row([
                        ft.Text(f"#{sid} — {st.get('title', '') or '—'}", size=12.5,
                               weight=ft.FontWeight.BOLD, color=T.INK, expand=True,
                               max_lines=1, overflow=ft.TextOverflow.ELLIPSIS),
                        ft.Text(f"{len(rows)}/{_TM_CT_MAX_ROWS} tasks", size=11,
                               color=T.INK_3),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    ft.Container(height=10),
                    ft.Column(row_blocks, spacing=0),
                    ft.Container(height=6),
                    add_btn,
                ], spacing=0),
                padding=14, margin=ft.Margin.only(bottom=10), bgcolor=T.CARD_2,
                border=ft.Border.all(1, T.BORDER_2), border_radius=T.R))

        _can_create = (bool(app._tm_ct_selected) and bool(app._tm_ct_user)
                      and not app._tm_ct_busy and not _ro)
        btn = green_btn(
            "Creating…" if app._tm_ct_busy else "Create child tasks",
            icon=ft.Icons.ADD_TASK,
            on_click=((lambda e: _create_tasks(app)) if _can_create else None))
        try:
            btn.opacity = 1.0 if _can_create else 0.45
        except Exception:
            pass

        parts = []
        if story_blocks:
            parts += [ft.Container(height=14),
                     field_label("Per-story details"),
                     ft.Container(height=6),
                     ft.Column(story_blocks, spacing=0)]
        parts += [ft.Container(height=14), btn]
        return ft.Column(parts, spacing=0)

    def _set_row_field(sid, idx, field):
        def _h(v_or_e):
            rows = app._tm_ct_rows.get(sid) or []
            if idx >= len(rows):
                return
            # _date_field's on_pick calls back with a plain "YYYY-MM-DD"
            # string; TextField's on_change calls back with an event object
            # (e.control.value) — same handler covers both call shapes.
            val = v_or_e if isinstance(v_or_e, str) else v_or_e.control.value
            rows[idx][field] = val
        return _h

    def _add_row(sid):
        def _h(e):
            rows = app._tm_ct_rows.setdefault(sid, [])
            if len(rows) < _TM_CT_MAX_ROWS:
                st = {s["id"]: s for s in app._tm_ct_stories}.get(sid, {})
                rows.append(_tm_default_row(st))
            _sync_ct_dynamic()
        return _h

    def _remove_row(sid, idx):
        def _h(e):
            rows = app._tm_ct_rows.get(sid) or []
            if len(rows) > 1 and idx < len(rows):
                rows.pop(idx)
            _sync_ct_dynamic()
        return _h

    def _sync_ct_dynamic():
        if getattr(app, "_tm_screen_epoch", 0) != _my_epoch:
            # Same staleness check as _sync_report_dynamic — a full render
            # elsewhere has already replaced this cell.
            app.render()
            return
        cell = _ct_dynamic_cell[0]
        if cell is None:
            return
        try:
            cell.content = _build_ct_dynamic()
            cell.update()
        except Exception:
            pass

    def _toggle_story(key, checked):
        s = set(app._tm_ct_selected)
        s.add(key) if checked else s.discard(key)
        app._tm_ct_selected = [x["id"] for x in app._tm_ct_stories if x["id"] in s]
        _sync_ct_dynamic()

    def _all_stories(checked):
        app._tm_ct_selected = [x["id"] for x in app._tm_ct_stories] if checked else []
        _sync_ct_dynamic()

    def _open_stories():
        app._tm_ct_open = not app._tm_ct_open

    def _set_ct_user(e):
        app._tm_ct_user = e.control.value or ""
        _sync_ct_dynamic()

    _ct_section_cell = [None]   # rebuilt-in-place: the WHOLE Section 2 card —
    # either the locked stack or the real interactive card — not just the
    # story-rows cell above. Needed as its own cell (distinct from
    # _ct_dynamic_cell) because the locked/unlocked decision itself has to
    # react to the scope toggle, which only ever does a PARTIAL update (see
    # _set_scope_mode above) — without this, flipping the toggle looked like
    # it did nothing to Section 2 until some unrelated full render caught up.

    def _build_ct_section():
        # Date range has no equivalent notion of "which stories" (that's a
        # sprint concept), so instead of quietly working off whatever sprint
        # happened to be picked last while in Sprint mode (the earlier
        # accepted trade-off), this whole section reads as fully LOCKED
        # whenever Date range is the active scope — same locked-card
        # treatment (skeleton rows + centered pill message) as Setup
        # screen's Step 3 "Task" card before Connect (main.py's
        # _task_locked), just with a message specific to this case. Read
        # fresh every call (not captured once) so this cell's own rebuild
        # always reflects whatever the toggle is set to AT THAT MOMENT.
        locked = (getattr(app, "_tm_scope_mode", "sprint") == "dates")

        if locked:
            return [card(ft.Stack([
                ft.Column([
                    sec_head("2", "Create child tasks",
                             ft.Row([ft.Icon(ft.Icons.LOCK_OUTLINE, size=13, color=T.INK_3),
                                    ft.Text("locked", size=11, color=T.INK_3,
                                           weight=ft.FontWeight.BOLD)], spacing=4, tight=True)),
                    ft.Container(height=14),
                    ft.Container(height=44, bgcolor=T.CARD_2,
                                border=ft.Border.all(1, T.BORDER), border_radius=T.R),
                    ft.Container(height=12),
                    ft.Container(height=44, bgcolor=T.CARD_2,
                                border=ft.Border.all(1, T.BORDER), border_radius=T.R),
                ], spacing=0),
                ft.Container(
                    ft.Container(
                        ft.Row([ft.Icon(ft.Icons.LOCK_OUTLINE, size=14, color=T.INK_2),
                               ft.Text("Pick a sprint to load stories", size=12,
                                      color=T.INK_2, weight=ft.FontWeight.BOLD)],
                              spacing=6, tight=True),
                        padding=ft.Padding.symmetric(vertical=14, horizontal=9),
                        bgcolor=T.CARD, border_radius=20, border=ft.Border.all(1, T.BORDER)),
                    alignment=ft.Alignment.CENTER, expand=True),
            ]), expand=True)]

        story_picker = (
            ft.Container(ft.Text("Loading stories…", color=T.INK_3, size=12), padding=10)
            if app._tm_ct_stories_loading else
            R._checkbox_multiselect(
                [(s["id"], f"#{s['id']} — {s['title']}") for s in app._tm_ct_stories],
                app._tm_ct_selected, _toggle_story, _all_stories,
                is_open=app._tm_ct_open, on_open=_open_stories,
                placeholder="Select user stories…",
                empty="No user stories found for this sprint.",
                page=app.page, app=app, sync_key="tm_stories", disabled=_ro))

        # Confirmed via the live diagnostics log (page.on_error -> diag_log) that
        # the grey-box crash was NOT caused by having two searchable_dropdown
        # instances on this screen — that was an unconfirmed second opinion, ruled
        # out once the real Flutter error ("WrapParentData is not a subtype of
        # FlexParentData") pointed at kpi_strip's wrap=True instead (see the note
        # there). Safe to restore this as searchable now.
        ct_user_dd = searchable_dropdown(
            value=app._tm_ct_user or None, options=_member_opts(), disabled=_ro,
            hint_text="Type to search a person…", text_size=13, border_color=T.BORDER,
            focused_border_color=T.VIOLET, border_radius=T.R,
            content_padding=ft.Padding.symmetric(vertical=8, horizontal=10),
            on_select=_set_ct_user)

        ct_dynamic = ft.Container(_build_ct_dynamic())
        _ct_dynamic_cell[0] = ct_dynamic

        section = [card(ft.Column([
            sec_head("2", "Create child tasks"),
            ft.Container(height=10),
            ft.Column([field_label("User stories", req=True), story_picker], spacing=6),
            ft.Container(height=14),
            ft.Column([field_label("Assign all to", req=True), hover_field(ct_user_dd)],
                      spacing=6),
            ct_dynamic,
        ], spacing=0))]

        banner2 = _status_banner(app._tm_ct_msg, app._tm_ct_busy)
        if banner2:
            section.append(banner2)

        ctres = app._tm_ct_result
        if ctres and ctres.get("errors"):
            section.append(ft.Container(
                ft.Column([ft.Text(e, size=12, color=T.RED) for e in ctres["errors"]], spacing=4),
                margin=ft.Margin.only(top=10), padding=12, bgcolor=T.RED_SOFT, border_radius=T.R))
        return section

    def _sync_ct_section():
        if getattr(app, "_tm_screen_epoch", 0) != _my_epoch:
            # Same staleness guard as _sync_report_dynamic/_sync_ct_dynamic —
            # a full render elsewhere has already replaced this cell.
            app.render()
            return
        cell = _ct_section_cell[0]
        if cell is None:
            return
        try:
            cell.content = ft.Column(_build_ct_section(), spacing=0)
            cell.update()
        except Exception:
            pass

    ct_section = ft.Container(ft.Column(_build_ct_section(), spacing=0))
    _ct_section_cell[0] = ct_section
    ct_body = [ct_section]

    body = ft.Column(
        report_body + [ft.Container(height=24)] + ct_body,
        spacing=0, scroll=ft.ScrollMode.AUTO, expand=True)

    return app.shell("Task Manager",
                     "Per-user task workload reports and bulk child-task creation",
                     body)
