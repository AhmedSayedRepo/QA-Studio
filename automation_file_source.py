"""Local CSV/XLSX test-case source for Automation.

The Automation generator consumes a provider-neutral ``stories_payload``
structure.  This module turns a simple, reviewable spreadsheet into that
structure without importing it into Azure DevOps or writing back to the file.
"""
from __future__ import annotations

import csv
import hashlib
import os
import re
from collections import OrderedDict


SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


def _text(value):
    return str(value or "").strip()


def _header(value):
    return re.sub(r"[^a-z0-9]+", " ", _text(value).lower()).strip()


def _field(row, *names):
    for name in names:
        value = row.get(name)
        if value:
            return _text(value)
    return ""


def _rows_from_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            raise ValueError("The CSV needs a header row.")
        return [{_header(k): _text(v) for k, v in row.items() if k is not None}
                for row in reader]


def _rows_from_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - bundled by requirements
        raise ValueError("Excel import is unavailable because openpyxl is missing.") from exc
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in book.worksheets:
            values = sheet.iter_rows(values_only=True)
            header = next(values, None)
            if not header or not any(_text(v) for v in header):
                continue
            keys = [_header(v) for v in header]
            return [{keys[i]: _text(value) for i, value in enumerate(row)
                     if i < len(keys) and keys[i]}
                    for row in values]
    finally:
        book.close()
    raise ValueError("The workbook needs a worksheet with a header row.")


def load_test_cases(path):
    """Read ordered test cases from a CSV/XLSX file.

    Required columns: ``Test Case Title`` and ``Action``.  ``Expected Result``
    is optional.  ``Test Case ID`` and ``Step`` are optional; when absent, the
    title and spreadsheet row order keep steps grouped and ordered.
    """
    path = os.path.abspath(_text(path))
    ext = os.path.splitext(path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError("Choose a CSV or XLSX test-case file.")
    if not os.path.isfile(path):
        raise ValueError("The selected test-case file is no longer available.")

    rows = _rows_from_csv(path) if ext == ".csv" else _rows_from_xlsx(path)
    grouped = OrderedDict()
    for position, row in enumerate(rows, start=1):
        title = _field(row, "test case title", "testcase title", "case title", "title")
        action = _field(row, "action", "step action", "test step", "steps")
        expected = _field(row, "expected result", "expected", "expected outcome", "result")
        case_id = _field(row, "test case id", "testcase id", "case id", "id")
        # Optional execution metadata lets a reviewed local workbook express
        # browser context operations (frame/window/dialog) without trying to
        # smuggle them through natural-language action text.  Azure DevOps
        # payloads simply omit these fields and retain their existing flow.
        metadata = {
            "element_type": _field(row, "element type", "control type"),
            "preferred_locator": _field(row, "preferred locator", "locator", "selector"),
            "frame_context": _field(row, "frame context", "frame"),
            "automation_note": _field(row, "automation note", "note"),
        }
        if not any((title, action, expected, case_id)):
            continue
        if not title:
            raise ValueError(f"Row {position}: Test Case Title is required.")
        if not action:
            raise ValueError(f"Row {position}: Action is required.")
        key = case_id or title
        case = grouped.setdefault(key, {"id": case_id or f"row-{position}",
                                        "title": title, "steps": []})
        step = {"index": len(case["steps"]) + 1,
                "action": action, "expected": expected}
        step.update({key: value for key, value in metadata.items() if value})
        case["steps"].append(step)
    if not grouped:
        raise ValueError("No test-case steps were found in the selected file.")
    return list(grouped.values())


def load_stories_payload(path):
    """Create one stable local source story containing all imported cases."""
    cases = load_test_cases(path)
    path = os.path.abspath(path)
    fingerprint = hashlib.sha256(path.encode("utf-8")).hexdigest()[:12]
    name = os.path.basename(path)
    for seq, case in enumerate(cases, start=1):
        # Avoid collisions with Azure work-item IDs while remaining stable across
        # repeated runs of the same file.
        case["id"] = f"file-{fingerprint}-{case['id']}-{seq}"
    return [{
        "story": {"id": f"file-{fingerprint}", "title": f"Imported file: {name}",
                  "criteria": "Test cases imported from a local file."},
        "test_cases": cases,
    }]
